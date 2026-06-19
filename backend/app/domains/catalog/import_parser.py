"""Catalog import parser for CSV and XLSX uploads.

CSV: reads UTF-8 and Windows-1251 (cp1251) automatically (tries UTF-8 first,
falls back to cp1251 on UnicodeDecodeError). XLSX: openpyxl, read-only +
data_only (cached formula values; an uncached formula reads as None).

Both formats funnel through the same row builder (`_build_row`) so a parsed
row is byte-for-byte identical regardless of source — the downstream pipeline
cannot tell CSV from XLSX. Column mapping is a case-insensitive header match
against the expected names below; the only required column is `brand_name`.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

EXPECTED_COLUMNS = [
    "brand_name",
    "inn",
    "manufacturer",
    "form",
    "dosage",
    "pack_size",
    "atx_code",
    "dispensing_type",
    "storage_type",
    "category",
    "base_price",
    "barcode",
]
DISPENSING = {"prescription", "otc", "special"}
STORAGE = {"normal", "cold", "frozen"}

# Shown to the user when they upload the legacy binary .xls format. Reused by
# the upload endpoint (HTTP 422) and the parser dispatcher so the wording stays
# in one place.
XLS_UNSUPPORTED_MESSAGE = "Поддерживаются файлы .xlsx и .csv; пересохраните файл как .xlsx"


def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode CSV; supported encodings: UTF-8 and cp1251")


def _normalize_header(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def _coerce(value: Any) -> str:
    """Normalize a raw cell (str from CSV, or str/int/float/None from XLSX) to a
    trimmed string. Integers and integral floats render without scientific
    notation so a numeric barcode like 4600123456789 stays digits, not 4.6E+12."""
    if value is None:
        return ""
    if isinstance(value, bool):  # bool is an int subclass — guard before int
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _build_row(raw_values: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    """Validate and type one row given canonical-column → raw-cell values.
    Returns (parsed, row_errors); identical logic for CSV and XLSX sources."""
    parsed: dict[str, Any] = {}
    row_errors: list[str] = []

    for canonical in EXPECTED_COLUMNS:
        raw_val = _coerce(raw_values.get(canonical))
        if canonical == "base_price":
            if raw_val:
                try:
                    parsed[canonical] = Decimal(raw_val.replace(",", "."))
                except (InvalidOperation, ValueError):
                    row_errors.append(f"base_price '{raw_val}' is not a number")
                    parsed[canonical] = None
            else:
                parsed[canonical] = None
        else:
            parsed[canonical] = raw_val or None

    # Required field
    if not parsed.get("brand_name"):
        row_errors.append("brand_name is required")

    # Enum-ish validation + defaults
    dt = parsed.get("dispensing_type")
    if dt and dt not in DISPENSING:
        row_errors.append(f"dispensing_type '{dt}' is invalid")
    elif not dt:
        parsed["dispensing_type"] = "otc"

    st = parsed.get("storage_type")
    if st and st not in STORAGE:
        row_errors.append(f"storage_type '{st}' is invalid")
    elif not st:
        parsed["storage_type"] = "normal"

    return parsed, row_errors


def parse_csv(raw: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (rows, errors) where rows have validated, typed values keyed
    by EXPECTED_COLUMNS, and errors look like `{"row": N, "messages": [...]}`."""
    text = _decode(raw)
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise ValueError("CSV has no header row")

    header_map = {_normalize_header(fn): fn for fn in reader.fieldnames if fn is not None}
    if "brand_name" not in header_map:
        raise ValueError("CSV must contain a 'brand_name' column")

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for line_no, row in enumerate(reader, start=2):  # start=2 → 1-based, +header
        raw_values = {
            canonical: (row.get(header_map[canonical]) if canonical in header_map else None)
            for canonical in EXPECTED_COLUMNS
        }
        parsed, row_errors = _build_row(raw_values)
        if row_errors:
            errors.append({"row": line_no, "messages": row_errors})
        else:
            rows.append(parsed)

    return rows, errors


def _is_blank(cell: Any) -> bool:
    return cell is None or (isinstance(cell, str) and cell.strip() == "")


def parse_xlsx(raw: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse the first worksheet of an XLSX file. First row is the header
    (case-insensitive match). Same (rows, errors) contract as parse_csv."""
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises many exception types on malformed input
        raise ValueError("Could not read XLSX file") from exc

    try:
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise ValueError("XLSX has no header row") from None

        col_idx: dict[str, int] = {}
        for idx, name in enumerate(header):
            if name is not None:
                col_idx[_normalize_header(str(name))] = idx
        if "brand_name" not in col_idx:
            raise ValueError("XLSX must contain a 'brand_name' column")

        rows: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []

        for line_no, row in enumerate(rows_iter, start=2):
            if row is None or all(_is_blank(c) for c in row):
                continue  # skip the trailing empty rows XLSX files carry
            raw_values = {
                canonical: (
                    row[col_idx[canonical]]
                    if canonical in col_idx and col_idx[canonical] < len(row)
                    else None
                )
                for canonical in EXPECTED_COLUMNS
            }
            parsed, row_errors = _build_row(raw_values)
            if row_errors:
                errors.append({"row": line_no, "messages": row_errors})
            else:
                rows.append(parsed)

        return rows, errors
    finally:
        wb.close()


def parse_import(
    raw: bytes, filename: str | None
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Pick the parser by file extension: .xlsx → openpyxl, .xls → friendly
    refusal, anything else → CSV (the historical default)."""
    name = (filename or "").strip().lower()
    if name.endswith(".xlsx"):
        return parse_xlsx(raw)
    if name.endswith(".xls"):
        raise ValueError(XLS_UNSUPPORTED_MESSAGE)
    return parse_csv(raw)
