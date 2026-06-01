"""Accountant sales-summary XLSX (arbitrary date range) — openpyxl.

Two sheets: «Сводка» (period totals + payment breakdown) and «Детализация»
(one row per receipt). Unlike the Z-report this is generated on the fly and
never cached — the range is arbitrary and the underlying data isn't frozen.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.domains.pos.schemas import SalesSummaryData

_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11)
_HEAD_FONT = Font(bold=True)

_STATUS_LABEL = {"sale": "Завершён", "return": "Возврат", "voided": "Отменён"}
_METHOD_LABEL = {
    "cash": "Наличные",
    "card": "Карта",
    "bank_transfer": "Перевод",
    "mixed": "Смешанная",
    "none": "—",
}


def _fmt(currency: str) -> str:
    return f'#,##0.00" {currency}"'


def render_sales_summary_xlsx(data: SalesSummaryData) -> bytes:
    wb = Workbook()
    _summary_sheet(wb.active, data)
    _detail_sheet(wb.create_sheet("Детализация"), data)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _summary_sheet(ws: Worksheet, data: SalesSummaryData) -> None:
    ws.title = "Сводка"
    money = _fmt(data.currency)
    row = 1

    def title(text: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    def section(text: str) -> None:
        nonlocal row
        row += 1
        ws.cell(row=row, column=1, value=text).font = _SECTION_FONT
        row += 1

    def kv(label: str, value: object, *, is_money: bool = False, bold: bool = False) -> None:
        nonlocal row
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        if bold:
            lc.font = _HEAD_FONT
            vc.font = _HEAD_FONT
        if is_money:
            vc.number_format = money
        row += 1

    title("СВОДНЫЙ ОТЧЁТ ПО ПРОДАЖАМ")
    kv("Период", f"{data.date_from:%d.%m.%Y} — {data.date_to:%d.%m.%Y}")
    if data.branch_name:
        kv("Филиал", data.branch_name)

    section("ИТОГИ")
    kv("Валовые продажи", float(data.gross_sales), is_money=True)
    kv("Скидки", float(data.total_discounts), is_money=True)
    kv("Возвраты", float(data.total_refunds), is_money=True)
    kv("Чистая сумма", float(data.net), is_money=True, bold=True)
    kv("Количество продаж", data.sales_count)
    kv("Количество возвратов", data.returns_count)

    bd = data.payment_breakdown
    section("РАЗБИВКА ПО ОПЛАТЕ")
    for label, amount in (
        ("Наличные", bd.cash),
        ("Карта", bd.card),
        ("Перевод", bd.bank_transfer),
        ("Смешанная", bd.mixed),
    ):
        kv(label, float(amount), is_money=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A2"


def _detail_sheet(ws: Worksheet, data: SalesSummaryData) -> None:
    money = _fmt(data.currency)
    headers = ["Дата", "Чек", "Кассир"]
    if data.show_branch_column:
        headers.append("Филиал")
    headers += ["Оплата", "Валовая", "Скидка", "Итог", "Статус"]

    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name).font = _HEAD_FONT

    money_cols = {headers.index(h) + 1 for h in ("Валовая", "Скидка", "Итог")}

    for i, r in enumerate(data.rows, start=2):
        values: list[object] = [
            r.completed_at.strftime("%d.%m.%Y %H:%M") if r.completed_at is not None else "—",
            r.receipt_number or "—",
            r.cashier_name or "—",
        ]
        if data.show_branch_column:
            values.append(r.branch_name or "—")
        values += [
            _METHOD_LABEL.get(r.payment_method, r.payment_method),
            float(r.gross),
            float(r.discount),
            float(r.net),
            _STATUS_LABEL.get(r.kind, r.kind),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            if col in money_cols:
                cell.number_format = money

    widths = [18, 12, 22]
    if data.show_branch_column:
        widths.append(20)
    widths += [12, 14, 12, 14, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
