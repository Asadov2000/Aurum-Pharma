"""Z-report XLSX (shift close) — written with openpyxl.

openpyxl reads *and* writes XLSX, so the Phase-3 catalog import can reuse this
single dependency instead of pulling a second (write-only) library.

Mirrors the receipt-PDF pattern: lazily render and cache in MinIO keyed by
shift_id. A closed shift is immutable, so the cached workbook is byte-stable on
repeat requests.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from minio.error import S3Error
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.core.storage import ensure_bucket, get_object, put_object
from app.domains.pos.schemas import ZReportData

_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11)
_LABEL_FONT = Font(bold=False)


def _fmt(currency: str) -> str:
    # e.g. '#,##0.00" TJS"'
    return f'#,##0.00" {currency}"'


def render_z_report_xlsx(data: ZReportData) -> bytes:
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Z-отчёт"
    money = _fmt(data.currency)

    row = 1

    def title(text: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = _TITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    def section(text: str) -> None:
        nonlocal row
        row += 1  # blank spacer line before each section
        ws.cell(row=row, column=1, value=text).font = _SECTION_FONT
        row += 1

    def kv(label: str, value: object, *, is_money: bool = False) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        if is_money:
            cell.number_format = money
            cell.alignment = Alignment(horizontal="right")
        row += 1

    title("Z-ОТЧЁТ ПО СМЕНЕ")

    for label, value in (
        ("Аптека", data.pharmacy_name),
        ("Филиал", data.branch_name),
        ("Касса", data.register_name),
        ("Смена (ID)", str(data.shift_id)),
        ("Кассир", data.cashier_name or "—"),
        ("Открыта", _dt(data.opened_at)),
        ("Закрыта", _dt(data.closed_at)),
    ):
        kv(label, value)

    section("ПРОДАЖИ")
    kv("Количество продаж", data.sales_count)
    kv("Валовая сумма", float(data.total_sales), is_money=True)
    kv("Скидки", float(data.total_discounts), is_money=True)

    section("ВОЗВРАТЫ")
    kv("Количество возвратов", data.returns_count)
    kv("Сумма возвратов", float(data.total_refunds), is_money=True)

    bd = data.payment_breakdown
    section("РАЗБИВКА ПО ОПЛАТЕ")
    for label, amount in (
        ("Наличные", bd.cash),
        ("Карта", bd.card),
        ("Перевод", bd.bank_transfer),
        ("Смешанная", bd.mixed),
    ):
        kv(label, float(amount), is_money=True)

    section("КАССА")
    kv("Начальная касса", float(data.initial_cash), is_money=True)
    for label, optional in (
        ("Ожидаемая касса", data.expected_cash),
        ("Фактическая касса", data.actual_cash),
        ("Расхождение", data.cash_difference),
    ):
        kv(label, _money_or_none(optional), is_money=optional is not None)
    kv("Причина расхождения", data.difference_reason or "—")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 26
    ws.freeze_panes = "A2"  # keep the title row visible while scrolling

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _dt(value: datetime | None) -> str:
    return value.strftime("%d.%m.%Y %H:%M") if value is not None else "—"


def _money_or_none(value: Decimal | None) -> object:
    return float(value) if value is not None else "—"


def get_or_render_z_report_xlsx(data: ZReportData) -> bytes:
    """Return the cached workbook for a closed shift, rendering and storing it on
    first request. Blocking — call from a worker thread."""
    key = f"zreports/{data.shift_id}.xlsx"
    try:
        bucket = ensure_bucket()
        return get_object(f"{bucket}/{key}")
    except S3Error:
        pass  # not generated yet

    xlsx = render_z_report_xlsx(data)
    put_object(
        object_name=key,
        data=xlsx,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return xlsx
