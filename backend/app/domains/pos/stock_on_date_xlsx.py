"""Stock-on-date XLSX (accountant) — openpyxl, one sheet.

Per-batch balance as of a date, reconstructed from the batch_movement ledger.
Generated on the fly (never cached): the date is arbitrary and the data is
live. Layout matches the other reports — bold header, TJS currency format,
frozen header row, sized columns.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.domains.pos.schemas import StockOnDateData
from app.domains.pos.xlsx_safe import xlsx_safe

_HEAD_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)


def _fmt(currency: str) -> str:
    return f'#,##0.00" {currency}"'


def render_stock_on_date_xlsx(data: StockOnDateData) -> bytes:
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Остатки"
    money = _fmt(data.currency)

    title = f"ОСТАТКИ НА {data.on_date:%d.%m.%Y}"
    if data.branch_name:
        title += f" · {data.branch_name}"
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT

    headers = ["Наименование", "МНН"]
    if data.show_branch_column:
        headers.append("Филиал")
    headers += ["Партия", "Срок годности", "Кол-во", "Закуп. цена", "Стоимость"]
    head_row = 3
    for col, name in enumerate(headers, start=1):
        ws.cell(row=head_row, column=col, value=name).font = _HEAD_FONT

    money_cols = {headers.index(h) + 1 for h in ("Закуп. цена", "Стоимость")}
    qty_col = headers.index("Кол-во") + 1
    value_col = headers.index("Стоимость") + 1

    r = head_row + 1
    for row in data.rows:
        values: list[object] = [row.name or "—", row.inn or "—"]
        if data.show_branch_column:
            values.append(row.branch_name or "—")
        values += [
            row.batch_number or "—",
            row.expires_at.strftime("%d.%m.%Y") if row.expires_at is not None else "—",
            float(row.qty),
            float(row.purchase_price),
            float(row.value),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=xlsx_safe(value))
            if col in money_cols:
                cell.number_format = money
        r += 1

    # Totals row.
    total_cell = ws.cell(row=r, column=1, value="ИТОГО")
    total_cell.font = _HEAD_FONT
    qty_cell = ws.cell(row=r, column=qty_col, value=float(data.total_qty))
    qty_cell.font = _HEAD_FONT
    val_cell = ws.cell(row=r, column=value_col, value=float(data.total_value))
    val_cell.font = _HEAD_FONT
    val_cell.number_format = money

    widths = [30, 22]
    if data.show_branch_column:
        widths.append(20)
    widths += [16, 14, 12, 14, 16]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=col).column_letter].width = width

    # Freeze everything above the first data row (title + header stay on screen).
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1).coordinate

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
