"""XLSX formula-injection guard: dangerous-leading-char strings are neutralised,
and a malicious drug name is written as text, not a formula."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.domains.pos.schemas import StockOnDateData, StockRow
from app.domains.pos.stock_on_date_xlsx import render_stock_on_date_xlsx
from app.domains.pos.xlsx_safe import xlsx_safe


@pytest.mark.parametrize(
    "value, expected",
    [
        ("=1+1", "'=1+1"),
        ("+1", "'+1"),
        ("-1", "'-1"),
        ("@x", "'@x"),
        ("\tx", "'\tx"),
        ("\rx", "'\rx"),
        ("Аспирин", "Аспирин"),  # normal text untouched
        ("", ""),
    ],
)
def test_xlsx_safe_strings(value: str, expected: str) -> None:
    assert xlsx_safe(value) == expected


def test_xlsx_safe_passes_through_non_strings() -> None:
    assert xlsx_safe(5) == 5
    assert xlsx_safe(Decimal("3.50")) == Decimal("3.50")
    assert xlsx_safe(None) is None


def test_stock_xlsx_neutralises_formula_in_drug_name() -> None:
    data = StockOnDateData(
        on_date=date(2026, 6, 1),
        branch_name=None,
        show_branch_column=False,
        currency="TJS",
        rows=[
            StockRow(
                name="=cmd|'/c calc'!A1",  # attacker-controlled brand_name
                inn=None,
                branch_name=None,
                batch_number="B-1",
                expires_at=date(2026, 12, 1),
                qty=Decimal("10"),
                purchase_price=Decimal("3.00"),
                value=Decimal("30.00"),
            )
        ],
        total_qty=Decimal("10"),
        total_value=Decimal("30.00"),
    )
    wb = load_workbook(BytesIO(render_stock_on_date_xlsx(data)))
    ws = wb.active
    # Find the name cell (column A, first data row after title+header).
    name_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "cmd" in cell.value:
                name_cell = cell
                break
    assert name_cell is not None
    assert name_cell.value == "'=cmd|'/c calc'!A1"  # prefixed → literal text
    assert name_cell.data_type == "s"  # string, NOT a formula ('f')
